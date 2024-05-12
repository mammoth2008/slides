# # 学生研究题目分配程序

# ## 功能

# 该程序的主要功能是从 Excel 文件中读取学生信息，并从 Markdown 文件中读取研究题目，然后将题目随机分配给学生，并将结果保存回 Excel 文件中。

# ## 使用方式

# 1. 准备文件：
#    - 确保学生信息文件为 Excel 格式（如 `students.xlsx`），且包含学生 ID 和姓名。
#    - 准备研究题目文件为 Markdown 格式（如 `il-task.md`），每行一个题目。

# 2. 运行程序：
#    - 需要 openpyxl 库，可通过 `pip install openpyxl` 安装。
#    - 在程序运行目录下执行 Python 脚本。

# 3. 选择文件：
#    - 程序会显示当前目录下的所有文件，用户需根据文件序号选择学生 Excel 文件和题目文件。

# 4. 题目分配：
#    - 程序会随机分配题目给每个学生，并将结果输出到控制台。
#    - 如果 Excel 文件中已有分配结果，程序会询问是否覆盖已有结果。

# 5. 重新分配：
#    - 程序会询问用户是否需要重新分配题目，用户可选择继续分配或结束程序。

# 6. 保存结果：
#    - 最终分配结果会自动保存回原始的 Excel 文件中。

# ## 运行逻辑

# 1. 文件读取：
#    - 列出当前目录下的所有文件，用户通过输入序号选择学生信息文件和题目文件。

# 2. 数据验证：
#    - 检查题目数量是否足够分配给所有学生，如果不足，则提示用户并退出程序。
#    - 检查 Excel 文件中是否已有分配结果，如果有，则提示用户是否覆盖。

# 3. 题目分配：
#    - 随机打乱题目列表，将题目逐一分配给学生，并输出分配结果。
#    - 每次输出分配结果后暂停 0.5 秒，方便查看。

# 4. 用户交互：
#    - 提供选项让用户决定是否重新分配题目，确保用户满意最终分配结果。

# 5. 结果保存：
#    - 将最终分配结果保存回 Excel 文件中，覆盖或新增题目列。

# ## 注意事项

# - 确保 Excel 文件格式正确，至少包含两列：学生 ID 和学生姓名。
# - 题目文件应为 Markdown 格式，每行一个题目。
# - 如果题目数量不足，请补充足够的题目以满足学生数量要求。

import openpyxl
import random
import os
import time

# 计算字符串的显示宽度（中文字符宽度为2，英文字符宽度为1）
def get_display_width(s):
    width = 0
    for char in s:
        if '\u4e00' <= char <= '\u9fff':  # 中文字符范围
            width += 2
        else:
            width += 1
    return width

# 获取当前目录下的所有文件
files = os.listdir('.')

# 展示所有文件并让用户选择
print("📂 当前目录下的文件列表:")
for i, file in enumerate(files):
    print(f"{i + 1}. {file}")

# 获取用户选择的文件
students_file_index = int(input("请选择学生 Excel 文件的序号: ")) - 1
topics_file_index = int(input("请选择题目文件的序号: ")) - 1

students_file = files[students_file_index]
topics_file = files[topics_file_index]

# 读取学生信息
wb = openpyxl.load_workbook(students_file)
sheet = wb.active

# 读取研究题目
with open(topics_file, 'r', encoding='utf-8') as file:
    topics = file.readlines()
    topics = [topic.strip() for topic in topics if topic.strip()]  # 去除空行和首尾空白

# 检查题目数量是否足够
num_students = sheet.max_row - 1  # 排除标题行
if len(topics) < num_students:
    print("⚠️ 题目数量不足，请补足题目后再运行程序。")
    exit()

# 检查是否已有分配结果
has_existing_assignments = any(sheet.cell(row=i, column=3).value for i in range(2, num_students + 2))

# 覆盖旧结果选项
if has_existing_assignments:
    overwrite = input("📝 发现已有分配结果，是否覆盖？(y/n): ").strip().lower()
    if overwrite != 'y':
        print("🚫 程序终止，未进行任何更改。")
        exit()

# 随机分配题目
def assign_topics():
    random.shuffle(topics)
    assigned_topics = topics[:num_students]
    
    max_name_length = max(get_display_width(sheet.cell(row=i, column=2).value) for i in range(2, num_students + 2))
    
    for i in range(2, num_students + 2):
        student_id = sheet.cell(row=i, column=1).value
        student_name = sheet.cell(row=i, column=2).value
        assigned_topic = assigned_topics[i - 2]
        sheet.cell(row=i, column=3, value=assigned_topic)  # 将题目写入第三列
        
        name_display_width = get_display_width(student_name)
        padding = ' ' * (max_name_length - name_display_width)
        
        print(f"{student_id}  {student_name}{padding}  题目: {assigned_topic}")
        time.sleep(0.5)  # 停顿 0.5 秒

# 进行初次分配
print("🔄 正在分配题目...")
assign_topics()

# 询问是否重新分配
while True:
    retry = input("🔄 是否重新分配题目？(y/n): ").strip().lower()
    if retry == 'y':
        print("🔄 正在重新分配题目...")
        assign_topics()
    else:
        break

# 保存分配结果到原有的 Excel 文件中
wb.save(students_file)
print(f"✅ 题目分配完成，结果已保存到 {students_file}")
